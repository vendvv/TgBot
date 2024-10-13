# import openpyxl module
import openpyxl
import datetime
from datetime import date

# Give the location of the file
path = "timetable.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute

sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row=1, column=1)

print(f"working with {cell_obj.value}")
cell_obj = sheet_obj.cell(row=3, column=1)
print(f"{cell_obj.value}")
MAX_ROW = 28
MIN_ROW = 5
MAX_COL = 7
MIN_COL = 2
today = date.today()
print(f"today is {today}")


def weekday():
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')
    year, month, day = map(int, current_date.split('-'))
    date = datetime.datetime(year, month, day) #0-sunday //6-saturday
    today = date.isoweekday()
    if today == 7:
        today = 0
    ################for testing purposes only##################
    #print("Enter the day of the week:")
    #today = int(input())
    ################for testing purposes only##################
    print(f"day of the week: {today}")
    return today

def cur_week(week_number, del_buf):
    cur_week_num = week_number - del_buf
    if cur_week_num > 4:
        cur_week_num -= 4
    #print(f"current week number: {cur_week_num}")
    return cur_week_num


labs_tmrrw = list() # содержит списки с: str, pair_num
#lab = list() #str, pair_num

def LR_checkNadd(str, start, end, num, addition): #не включает пробелы!!!!
    if str[start:end] == "(ЛР)":
        str = addition + str
        lab = list()
        lab.append(str)
        lab.append(num)
        labs_tmrrw.append(lab)

def PZ_checkNadd(str, start, end, num, addition): #не включает пробелы!!!!  #для оуис
    if str[start:end] == "(ПЗ)":
        str = addition + str
        lab = list()
        lab.append(str)
        lab.append(num)
        labs_tmrrw.append(lab)

def get_tmrrw_tt(week_num):
    col = weekday()+2
    print(f"col number {col}")
    pair_num = 0
    for i in range(MIN_ROW+week_num-1, MAX_ROW, 4):
        pair_num += 1
        cell_obj = sheet_obj.cell(row=i, column=col)
        if type(cell_obj.value) != type(None):
            print(f"содержимое ячейки: <{cell_obj.value}, длина строки = {len(cell_obj.value)}>")
            if len(cell_obj.value) > 70:  # проверка на 2 группы
                buff = cell_obj.value[0:2]  # 1)
                for i in range(3, len(cell_obj.value)):
                    if cell_obj.value[i] == buff[0]:
                        if cell_obj.value[i + 1] == buff[1]:
                            lab = list()
                            lab.append(cell_obj.value[3:i-2])
                            lab.append(pair_num)
                            labs_tmrrw.append(lab)
                            lab2 = list()
                            lab2.append(cell_obj.value[i + 3:-1])
                            lab2.append(pair_num+ 0.1)
                            labs_tmrrw.append(lab2)
            else:  # newStr = cell_obj[
                # tmpstr = cell_obj.value[4:7]
                # print(f"cell_obj.value[4,7] = <{tmpstr}>")
                #print(f"cell_obj.value[4:7] = <{cell_obj.value[4:7]}>")
                if cell_obj.value[3:6] == "1п." or cell_obj.value[3:6] == "2п.":  # проверка всего остального
                    newStr = cell_obj.value[7:]
                    add = cell_obj.value[3:6]
                else:
                    newStr = cell_obj.value[3:]  # строка, начинающаяся с названия предмета
                    add = ''
                print(f"new string <{newStr}>, newStr[0:9] = <{newStr[0:8]}>")
                if newStr[0:2] == "БД" or newStr[0:2] == "КС":
                    LR_checkNadd(newStr, 3, 7, pair_num, add)
                elif newStr[0:5] == "ООПиП":
                    LR_checkNadd(newStr, 6, 10, pair_num, add)
                elif newStr[0:3] == "СЯП" or newStr[0:8] == "ОБА (ЛР)":
                    LR_checkNadd(newStr, 4, 8, pair_num, add)
                elif newStr[0:8] == "ОБА (ПЗ)":
                    print(f"newStr[4:8] = {newStr[4:8]}")
                    PZ_checkNadd(newStr, 4, 8, pair_num, add)
                elif newStr[0:4] == "ОУИС":
                    PZ_checkNadd(newStr, 5, 9, pair_num, add)

def start_xls():
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')
    year, month, day = map(int, current_date.split('-'))
    print(current_date)
    # Вывод дня недели
    # print(f"the weekday = {date.weekday()}")
    current_date = datetime.datetime.now()
    week_number = current_date.isocalendar()[1]
    print("enter current week number: ")
    cur_week_num = int(input())

    del_buf = week_number - cur_week_num  # удаляемый буфер
    # from now cur_week_num = week_number - del_buf

    week_num = cur_week(week_number, del_buf)
    print(f"current week number: {week_num}")

    get_tmrrw_tt(week_num)
    print(labs_tmrrw)


